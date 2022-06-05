from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class ExcelManipulate:
    def __init__(self, path:str, senderInfoList:list):
        self.path = path
        self.senderInfoList = senderInfoList
        self.wb = load_workbook(path)
        self.ws = self.wb.active
        self.writeBackColumnIndex = "K" #用來指定到時候資訊要寫回Excel的哪一個欄位
    # path 是要讀取的excel位址
    def read7_11Excel(self):

        resultLists = [] # 用來承裝所有在Excel當中讀到每一列資料
        columeIndex = (2, 3, 9, 8, 7) # 只選取Excel中特定欄位的資料
        maxRow = self.ws.max_row

        for row in range(1,maxRow+1):
            infolist = [1000] # 第一個是固定的包裹價值，直接寫死1000
            infolist = infolist + self.senderInfoList
            for colume in columeIndex:
                target = get_column_letter(colume) + str(row)
                result = self.ws[target].value
                infolist.append(result)

            resultLists.append(infolist)

        return resultLists

    # 原始result 會是 10個的list 但經過處理後，這邊傳進來的list 會是
    def writeBackToExcel(self, rowIndex:int ,resultString:str):
        self.ws[self.writeBackColumnIndex+str(rowIndex)].value = resultString
        self.wb.save(self.path)

